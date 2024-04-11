import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib.pagesizes import letter
import pyperclip

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    if file_path:
        entry_file.delete(0, tk.END)
        entry_file.insert(0, file_path)

def extract_data():
    file_path = entry_file.get()
    student_id = entry_id.get()
    if not file_path:
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == student_id:
            student_data = [cell.value for cell in row]
            generate_pdf(student_data)
            break
    else:
        messagebox.showwarning("Error", "Student ID not found.")

def generate_pdf(student_data):
    pdf_file = f"student_{student_data[0]}.pdf"
    c = canvas.Canvas(pdf_file, pagesize=letter)

    data = [['Student ID', 'Name', 'Subject', 'Marks'], student_data]
    table = Table(data, colWidths=[100, 100, 100, 100])

    style = TableStyle([('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('BACKGROUND', (0, 0), (-1, 0), (0.8, 0.8, 0.8))])

    table.setStyle(style)

    table.wrapOn(c, 400, 300)
    table.drawOn(c, 80, 700)

    c.save()
    messagebox.showinfo("PDF Generated", f"PDF generated: {pdf_file}")

def copy_row_to_clipboard():
    file_path = entry_file.get()
    student_id = entry_id.get()
    if not file_path:
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == student_id:
            student_data = [str(cell.value) for cell in row]
            row_data = "\t".join(student_data)
            pyperclip.copy(row_data)
            messagebox.showinfo("Copy Row", "Student data copied to clipboard.")
            break
    else:
        messagebox.showwarning("Error", "Student ID not found.")

window = tk.Tk()
window.title("Student Data Extractor")
window.geometry("400x350")
window.resizable(False, False)

label_file = tk.Label(window, text="Select Excel File:")
label_file.pack(pady=10)

entry_file = tk.Entry(window, width=50)
entry_file.pack(pady=5)

button_browse = tk.Button(window, text="Browse", command=select_file)
button_browse.pack(pady=5)

label_id = tk.Label(window, text="Enter Student ID:")
label_id.pack()

entry_id = tk.Entry(window)
entry_id.pack(pady=5)

button_extract = tk.Button(window, text="Extract Data", command=extract_data)
button_extract.pack(pady=10)

button_copy = tk.Button(window, text="Copy Row to Clipboard", command=copy_row_to_clipboard)
button_copy.pack(pady=10)

window.mainloop()
